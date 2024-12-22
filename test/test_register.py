import time

from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook, Workbook
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# File paths
input_file = "data/test_data_register.xlsx"
output_file = "test_results_register.xlsx"


service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)
driver.set_window_size(1920, 1080)


# Read data from Excel
def read_excel(sheet_name):
    workbook = load_workbook(input_file)
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    workbook.close()
    return data


# Get error messages from the page
def get_error_messages():
    messages = []
    error_elements = driver.find_elements(
        By.XPATH, "//*[contains(@class, 'error') or contains(@class, 'alert')]"
    )
    for element in error_elements:
        messages.append(element.text)
    return messages


# Get toast message from the page
def get_toast_message():
    try:
        toast_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "Toastify__toast-body"))
        )
        return toast_element.text
    except Exception:
        return "Toast message not found"


# Test Sign Up
def test_sign_up():
    results = []
    data = read_excel("RegisterData")
    for i, (
        name,
        email,
        password,
        confirmPassword,
        expected_message,  # Expected toast message
    ) in enumerate(data, start=1):
        try:
            driver.get("http://localhost:3000/auth/register")
            time.sleep(5)

            # Fill fields if data is provided
            if name:
                name_input = driver.find_element(By.ID, "name")
                name_input.clear()
                name_input.send_keys(name)
            if email:
                email_input = driver.find_element(By.ID, "email")
                email_input.clear()
                email_input.send_keys(email)
            if password:
                password_input = driver.find_element(By.ID, "password")
                password_input.clear()
                password_input.send_keys(password)
            if confirmPassword:
                confirmPassword_input = driver.find_element(By.ID, "confirm_password")
                confirmPassword_input.clear()
                confirmPassword_input.send_keys(confirmPassword)

            time.sleep(2)

            # Click register button
            register_button = driver.find_element(
                By.XPATH, '//*[@class="btn btn-primary border-0 w-100 mt-2"]'
            )
            register_button.click()
            time.sleep(5)

            # Get actual toast message
            actual_message = get_toast_message()
            if expected_message in actual_message:
                results.append(
                    (
                        f"Sign Up Test {i}",
                        "Passed",
                        "User signed up successfully",
                        actual_message,
                    )
                )
            else:
                results.append(
                    (
                        f"Sign Up Test {i}",
                        "Failed",
                        "Sign Up failed or unexpected toast message",
                        "Invalid input field",
                    )
                )

        except Exception as e:
            results.append((f"Sign Up Test {i}", "Error", str(e), "N/A"))

        # Wait for a brief moment before starting the next test
        time.sleep(3)

    return results


def get_fill_color(status):
    if status == "Error":
        return PatternFill(
            start_color="FF0000", end_color="FF0000", fill_type="solid"
        )  # Red
    elif status == "Failed":
        return PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )  # Yellow
    elif status == "Passed":
        return PatternFill(
            start_color="00FF00", end_color="00FF00", fill_type="solid"
        )  # Green
    return None


# Write results to specified sheet in Excel with colored status cells
def write_results(workbook, sheet_name, results):
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(["Test Case", "Status", "Message", "Actual Message"])

    for result in results:
        row = sheet.max_row + 1
        sheet.append(result)

        # Apply fill color to the Status cell based on the result's status
        status_fill = get_fill_color(result[1])
        if status_fill:
            sheet.cell(row=row, column=2).fill = status_fill


# Run tests and save results to Excel
workbook = Workbook()
workbook.remove(workbook.active)  # Remove default sheet

sign_up_results = test_sign_up()

write_results(workbook, "SignUpResults", sign_up_results)

workbook.save(output_file)

# Close WebDriver
time.sleep(2)
driver.quit()
# Đăng ký thành công. Kiểm tra email để kích hoạt tài khoản
# Toastify__toast-container Toastify__toast-container--top-right
# Toastify__toast Toastify__toast-theme--colored Toastify__toast--error Toastify__toast--close-on-click
