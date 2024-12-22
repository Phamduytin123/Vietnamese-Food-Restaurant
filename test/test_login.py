import time

from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook, Workbook
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options

# File paths
input_file = "data/test_data_login.xlsx"
output_file = "test_results_login.xlsx"

service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)
driver.set_window_size(1920, 1080)


# Read data from Excel
def read_excel(sheet_name):
    workbook = load_workbook(input_file)
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    workbook.close()
    return data


# Get error messages from the page
def get_error_messages():
    messages = []
    error_elements = driver.find_elements(
        By.XPATH, "//*[contains(@class, 'error') or contains(@class, 'alert')]"
    )
    for element in error_elements:
        messages.append(element.text)
    return messages


# Test Sign Up
def focus_on_page():
    try:
        # Click on the body of the page or any specific element to focus on the page tab
        driver.find_element(By.TAG_NAME, "body").click()
    except Exception as e:
        print(f"Error focusing on page: {str(e)}")


# Test Sign Up


# Test Login
def test_login():
    results = []
    data = read_excel("LoginData")
    for i, (email, password, expected_link) in enumerate(data, start=1):
        try:
            driver.get("http://localhost:3000/auth/login")
            time.sleep(5)

            # Focus on the page tab
            focus_on_page()

            if email:
                email_input = driver.find_element(By.ID, "email")
                email_input.clear()
                email_input.send_keys(email)
            if password:
                password_input = driver.find_element(By.ID, "password")
                password_input.clear()
                password_input.send_keys(password)
            print("oke1")

            login_button = driver.find_element(
                By.XPATH, '//*[@class="btn btn-primary border-0 w-100 mt-2"]'
            )
            login_button.click()
            print("oke2")
            time.sleep(5)

            # Increase the wait time to allow more time for login processing
            time.sleep(5)  # Increased sleep time for longer wait (5 seconds)

            actual_link = driver.current_url  # Get the actual link after login attempt

            if actual_link == expected_link:
                results.append(
                    (
                        f"Login Test {i}",
                        email,
                        password,
                        "Passed",
                        "Login successful",
                        actual_link,
                    )
                )
            else:
                results.append(
                    (
                        f"Login Test {i}",
                        email,
                        password,
                        "Failed",
                        "Login failed - Invalid credentials",
                        actual_link,
                    )
                )
            print("oke3")
        except Exception as e:
            actual_link = driver.current_url
            results.append(
                (f"Login Test {i}", email, password, "Error", str(e), actual_link)
            )
        print(results)
    return results


def get_fill_color(status):
    if status == "Error":
        return PatternFill(
            start_color="FF0000", end_color="FF0000", fill_type="solid"
        )  # Red
    elif status == "Failed":
        return PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )  # Yellow
    elif status == "Passed":
        return PatternFill(
            start_color="00FF00", end_color="00FF00", fill_type="solid"
        )  # Green
    return None


# Write results to specified sheet in Excel with colored status cells
def write_results(workbook, sheet_name, results):
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(["Test Case", "Email", "Password", "Status", "Message", "Actual Link"])

    for result in results:
        row = sheet.max_row + 1
        sheet.append(result)

        # Apply fill color to the Status cell based on the result's status
        status_fill = get_fill_color(result[1])
        if status_fill:
            sheet.cell(row=row, column=2).fill = status_fill


# Run tests and save results to Excel
workbook = Workbook()
workbook.remove(workbook.active)  # Remove default sheet

login_results = test_login()

write_results(workbook, "LoginResults", login_results)

workbook.save(output_file)

# Close WebDriver
time.sleep(2)
driver.quit()
